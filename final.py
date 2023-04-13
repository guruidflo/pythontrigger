import requests
import json
import pandas as pd
from pathlib import Path
from datetime import datetime
from datetime import timedelta
from datetime import date
from openpyxl import Workbook
import firebase_admin
from firebase_admin import credentials, firestore
import re
import pandas as pd
import numpy as np
from openpyxl import load_workbook
import openpyxl
import smtplib
from email.message import EmailMessage
import datetime
from datetime import datetime
from datetime import timedelta
from datetime import date


date_str = datetime.now().strftime('%Y-%m-%d')
date_input = datetime.now().strftime('%Y%m%d')
json_file_name = 'file-'+date_str+'.json'
xlsx_file_name =  'file-'+date_str+'.xlsx'  

cred = credentials.Certificate("fabtrakr-gsheet-token-firebase-adminsdk-jcky7-dc35d5ecac.json")
firebase_admin.initialize_app(cred)

db = firestore.client()  
collection = db.collection('tokens') 
docs = collection.get()
token = docs[0].to_dict()['token']

url = "https://api.fabtrakr.com/analytics/sewing/overView"
parameters = {
  "startDate" : date_input,
  "endDate" : date_input
}
payload={}
headers = {
  'version': '7.0',
  'Authorization': 'Bearer ' + str(token) 
  }

response = requests.request("GET", url,params=parameters, headers=headers, data=payload)
jsonData = json.loads(response.text)
with open(json_file_name, "w") as outfile:

    json.dump(jsonData, outfile)
p=Path(json_file_name)

with p.open('r', encoding='utf-8') as f:
   data = json.loads(f.read())

df = pd.read_json(json_file_name)
df.drop(['locationWorkDays','batchUnique','batchWorkDays'], axis=1, inplace = True)

writer = pd.ExcelWriter(xlsx_file_name)
df.to_excel(writer)

writer.save()

FilePath = (xlsx_file_name)
ExcelWorkbook = load_workbook(FilePath)
writer = pd.ExcelWriter(FilePath, engine = 'openpyxl')
writer.book = ExcelWorkbook

DataSample1= [['IDD1'], ['IDU1'],['IDU3'],['IDU4'],['IDU5'],['IDU8'],['IDU9P1'],['NJU1P1'],['NJU1P2'],['NJU2'],['NJU3']]

SimpleDataFrame1=pd.DataFrame(data=DataSample1, columns=['location'])

print(SimpleDataFrame1)

SimpleDataFrame1.to_excel(writer, sheet_name = 'Summary')
writer.save()


wb = openpyxl.load_workbook(xlsx_file_name)
Sheet = wb.get_sheet_by_name('Summary')
Sheet1 = wb.get_sheet_by_name('Sheet1')
Sheet['C1'] = 'batchTargetQuantity'
Sheet['D1'] = 'outputQuantity'
Sheet['E1'] = 'sewingSamProduced'
Sheet['F1'] = 'sewingMachineMinutes'
Sheet['G1'] = 'globalSamProduced'
Sheet['H1'] = 'globalMachineMinutes'
Sheet['I1'] = 'totalWorkStation'
Sheet['J1'] = 'numberOfMachines'
Sheet['K1'] = 'DHUPassedQuantity'
Sheet['L1'] = 'DHUReworkedQuantity'
Sheet['M1'] = 'DHURejectedQuantity'
Sheet['N1'] = 'DHU%'
Sheet['O1'] = 'SEWING EFFICIENCY'
Sheet['P1'] = 'GLOBAL EFFICIENCY'
Sheet1['O1'] = 'DHU%'
Sheet1['P1'] = 'SEWING EFFICIENCY'
Sheet1['Q1'] = 'GLOBAL EFFICIENCY'
Sheet['C2'] = '=SUMIF(Sheet1!$B$2:$B$55,$B2,Sheet1!D$2:D$55)'
Sheet['C3'] = '=SUMIF(Sheet1!$B$2:$B$55,$B3,Sheet1!D$2:D$55)'
Sheet['C4'] = '=SUMIF(Sheet1!$B$2:$B$55,$B4,Sheet1!D$2:D$55)'
Sheet['C5'] = '=SUMIF(Sheet1!$B$2:$B$55,$B5,Sheet1!D$2:D$55)'
Sheet['C6'] = '=SUMIF(Sheet1!$B$2:$B$55,$B6,Sheet1!D$2:D$55)'
Sheet['C7'] = '=SUMIF(Sheet1!$B$2:$B$55,$B7,Sheet1!D$2:D$55)'
Sheet['C8'] = '=SUMIF(Sheet1!$B$2:$B$55,$B8,Sheet1!D$2:D$55)'
Sheet['C9'] = '=SUMIF(Sheet1!$B$2:$B$55,$B9,Sheet1!D$2:D$55)'
Sheet['C10'] = '=SUMIF(Sheet1!$B$2:$B$55,$B10,Sheet1!D$2:D$55)'
Sheet['C11'] = '=SUMIF(Sheet1!$B$2:$B$55,$B11,Sheet1!D$2:D$55)'
Sheet['C12'] = '=SUMIF(Sheet1!$B$2:$B$55,$B12,Sheet1!D$2:D$55)'
Sheet['D2'] = '=SUMIF(Sheet1!$B$2:$B$55,$B2,Sheet1!E$2:E$55)'
Sheet['D3'] = '=SUMIF(Sheet1!$B$2:$B$55,$B3,Sheet1!E$2:E$55)'
Sheet['D4'] = '=SUMIF(Sheet1!$B$2:$B$55,$B4,Sheet1!E$2:E$55)'
Sheet['D5'] = '=SUMIF(Sheet1!$B$2:$B$55,$B5,Sheet1!E$2:E$55)'
Sheet['D6'] = '=SUMIF(Sheet1!$B$2:$B$55,$B6,Sheet1!E$2:E$55)'
Sheet['D7'] = '=SUMIF(Sheet1!$B$2:$B$55,$B7,Sheet1!E$2:E$55)'
Sheet['D8'] = '=SUMIF(Sheet1!$B$2:$B$55,$B8,Sheet1!E$2:E$55)'
Sheet['D9'] = '=SUMIF(Sheet1!$B$2:$B$55,$B9,Sheet1!E$2:E$55)'
Sheet['D10'] = '=SUMIF(Sheet1!$B$2:$B$55,$B10,Sheet1!E$2:E$55)'
Sheet['D11'] = '=SUMIF(Sheet1!$B$2:$B$55,$B11,Sheet1!E$2:E$55)'
Sheet['D12'] = '=SUMIF(Sheet1!$B$2:$B$55,$B12,Sheet1!E$2:E$55)'
Sheet['E2'] = '=SUMIF(Sheet1!$B$2:$B$55,$B2,Sheet1!F$2:F$55)'
Sheet['E3'] = '=SUMIF(Sheet1!$B$2:$B$55,$B3,Sheet1!F$2:F$55)'
Sheet['E4'] = '=SUMIF(Sheet1!$B$2:$B$55,$B4,Sheet1!F$2:F$55)'
Sheet['E5'] = '=SUMIF(Sheet1!$B$2:$B$55,$B5,Sheet1!F$2:F$55)'
Sheet['E6'] = '=SUMIF(Sheet1!$B$2:$B$55,$B6,Sheet1!F$2:F$55)'
Sheet['E7'] = '=SUMIF(Sheet1!$B$2:$B$55,$B7,Sheet1!F$2:F$55)'
Sheet['E8'] = '=SUMIF(Sheet1!$B$2:$B$55,$B8,Sheet1!F$2:F$55)'
Sheet['E9'] = '=SUMIF(Sheet1!$B$2:$B$55,$B9,Sheet1!F$2:F$55)'
Sheet['E10'] = '=SUMIF(Sheet1!$B$2:$B$55,$B10,Sheet1!F$2:F$55)'
Sheet['E11'] = '=SUMIF(Sheet1!$B$2:$B$55,$B11,Sheet1!F$2:F$55)'
Sheet['E12'] = '=SUMIF(Sheet1!$B$2:$B$55,$B12,Sheet1!F$2:F$55)'
Sheet['F2'] = '=SUMIF(Sheet1!$B$2:$B$55,$B2,Sheet1!G$2:G$55)'
Sheet['F3'] = '=SUMIF(Sheet1!$B$2:$B$55,$B3,Sheet1!G$2:G$55)'
Sheet['F4'] = '=SUMIF(Sheet1!$B$2:$B$55,$B4,Sheet1!G$2:G$55)'
Sheet['F5'] = '=SUMIF(Sheet1!$B$2:$B$55,$B5,Sheet1!G$2:G$55)'
Sheet['F6'] = '=SUMIF(Sheet1!$B$2:$B$55,$B6,Sheet1!G$2:G$55)'
Sheet['F7'] = '=SUMIF(Sheet1!$B$2:$B$55,$B7,Sheet1!G$2:G$55)'
Sheet['F8'] = '=SUMIF(Sheet1!$B$2:$B$55,$B8,Sheet1!G$2:G$55)'
Sheet['F9'] = '=SUMIF(Sheet1!$B$2:$B$55,$B9,Sheet1!G$2:G$55)'
Sheet['F10'] = '=SUMIF(Sheet1!$B$2:$B$55,$B10,Sheet1!G$2:G$55)'
Sheet['F11'] = '=SUMIF(Sheet1!$B$2:$B$55,$B11,Sheet1!G$2:G$55)'
Sheet['F12'] = '=SUMIF(Sheet1!$B$2:$B$55,$B12,Sheet1!G$2:G$55)'
Sheet['G2'] = '=SUMIF(Sheet1!$B$2:$B$55,$B2,Sheet1!H$2:H$55)'
Sheet['G3'] = '=SUMIF(Sheet1!$B$2:$B$55,$B3,Sheet1!H$2:H$55)'
Sheet['G4'] = '=SUMIF(Sheet1!$B$2:$B$55,$B4,Sheet1!H$2:H$55)'
Sheet['G5'] = '=SUMIF(Sheet1!$B$2:$B$55,$B5,Sheet1!H$2:H$55)'
Sheet['G6'] = '=SUMIF(Sheet1!$B$2:$B$55,$B6,Sheet1!H$2:H$55)'
Sheet['G7'] = '=SUMIF(Sheet1!$B$2:$B$55,$B7,Sheet1!H$2:H$55)'
Sheet['G8'] = '=SUMIF(Sheet1!$B$2:$B$55,$B8,Sheet1!H$2:H$55)'
Sheet['G9'] = '=SUMIF(Sheet1!$B$2:$B$55,$B9,Sheet1!H$2:H$55)'
Sheet['G10'] = '=SUMIF(Sheet1!$B$2:$B$55,$B10,Sheet1!H$2:H$55)'
Sheet['G11'] = '=SUMIF(Sheet1!$B$2:$B$55,$B11,Sheet1!H$2:H$55)'
Sheet['G12'] = '=SUMIF(Sheet1!$B$2:$B$55,$B12,Sheet1!H$2:H$55)'
Sheet['H2'] = '=SUMIF(Sheet1!$B$2:$B$55,$B2,Sheet1!I$2:I$55)'
Sheet['H3'] = '=SUMIF(Sheet1!$B$2:$B$55,$B3,Sheet1!I$2:I$55)'
Sheet['H4'] = '=SUMIF(Sheet1!$B$2:$B$55,$B4,Sheet1!I$2:I$55)'
Sheet['H5'] = '=SUMIF(Sheet1!$B$2:$B$55,$B5,Sheet1!I$2:I$55)'
Sheet['H6'] = '=SUMIF(Sheet1!$B$2:$B$55,$B6,Sheet1!I$2:I$55)'
Sheet['H7'] = '=SUMIF(Sheet1!$B$2:$B$55,$B7,Sheet1!I$2:I$55)'
Sheet['H8'] = '=SUMIF(Sheet1!$B$2:$B$55,$B8,Sheet1!I$2:I$55)'
Sheet['H9'] = '=SUMIF(Sheet1!$B$2:$B$55,$B9,Sheet1!I$2:I$55)'
Sheet['H10'] = '=SUMIF(Sheet1!$B$2:$B$55,$B10,Sheet1!I$2:I$55)'
Sheet['H11'] = '=SUMIF(Sheet1!$B$2:$B$55,$B11,Sheet1!I$2:I$55)'
Sheet['H12'] = '=SUMIF(Sheet1!$B$2:$B$55,$B12,Sheet1!I$2:I$55)'
Sheet['I2'] = '=SUMIF(Sheet1!$B$2:$B$55,$B2,Sheet1!J$2:J$55)'
Sheet['I3'] = '=SUMIF(Sheet1!$B$2:$B$55,$B3,Sheet1!J$2:J$55)'
Sheet['I4'] = '=SUMIF(Sheet1!$B$2:$B$55,$B4,Sheet1!J$2:J$55)'
Sheet['I5'] = '=SUMIF(Sheet1!$B$2:$B$55,$B5,Sheet1!J$2:J$55)'
Sheet['I6'] = '=SUMIF(Sheet1!$B$2:$B$55,$B6,Sheet1!J$2:J$55)'
Sheet['I7'] = '=SUMIF(Sheet1!$B$2:$B$55,$B7,Sheet1!J$2:J$55)'
Sheet['I8'] = '=SUMIF(Sheet1!$B$2:$B$55,$B8,Sheet1!J$2:J$55)'
Sheet['I9'] = '=SUMIF(Sheet1!$B$2:$B$55,$B9,Sheet1!J$2:J$55)'
Sheet['I10'] = '=SUMIF(Sheet1!$B$2:$B$55,$B10,Sheet1!J$2:J$55)'
Sheet['I11'] = '=SUMIF(Sheet1!$B$2:$B$55,$B11,Sheet1!J$2:J$55)'
Sheet['I12'] = '=SUMIF(Sheet1!$B$2:$B$55,$B12,Sheet1!J$2:J$55)'
Sheet['J2'] = '=SUMIF(Sheet1!$B$2:$B$55,$B2,Sheet1!K$2:K$55)'
Sheet['J3'] = '=SUMIF(Sheet1!$B$2:$B$55,$B3,Sheet1!K$2:K$55)'
Sheet['J4'] = '=SUMIF(Sheet1!$B$2:$B$55,$B4,Sheet1!K$2:K$55)'
Sheet['J5'] = '=SUMIF(Sheet1!$B$2:$B$55,$B5,Sheet1!K$2:K$55)'
Sheet['J6'] = '=SUMIF(Sheet1!$B$2:$B$55,$B6,Sheet1!K$2:K$55)'
Sheet['J7'] = '=SUMIF(Sheet1!$B$2:$B$55,$B7,Sheet1!K$2:K$55)'
Sheet['J8'] = '=SUMIF(Sheet1!$B$2:$B$55,$B8,Sheet1!K$2:K$55)'
Sheet['J9'] = '=SUMIF(Sheet1!$B$2:$B$55,$B9,Sheet1!K$2:K$55)'
Sheet['J10'] = '=SUMIF(Sheet1!$B$2:$B$55,$B10,Sheet1!K$2:K$55)'
Sheet['J11'] = '=SUMIF(Sheet1!$B$2:$B$55,$B11,Sheet1!K$2:K$55)'
Sheet['J12'] = '=SUMIF(Sheet1!$B$2:$B$55,$B12,Sheet1!K$2:K$55)'
Sheet['K2'] = '=SUMIF(Sheet1!$B$2:$B$55,$B2,Sheet1!L$2:L$55)'
Sheet['K3'] = '=SUMIF(Sheet1!$B$2:$B$55,$B3,Sheet1!L$2:L$55)'
Sheet['K4'] = '=SUMIF(Sheet1!$B$2:$B$55,$B4,Sheet1!L$2:L$55)'
Sheet['K5'] = '=SUMIF(Sheet1!$B$2:$B$55,$B5,Sheet1!L$2:L$55)'
Sheet['K6'] = '=SUMIF(Sheet1!$B$2:$B$55,$B6,Sheet1!L$2:L$55)'
Sheet['K7'] = '=SUMIF(Sheet1!$B$2:$B$55,$B7,Sheet1!L$2:L$55)'
Sheet['K8'] = '=SUMIF(Sheet1!$B$2:$B$55,$B8,Sheet1!L$2:L$55)'
Sheet['K9'] = '=SUMIF(Sheet1!$B$2:$B$55,$B9,Sheet1!L$2:L$55)'
Sheet['K10'] = '=SUMIF(Sheet1!$B$2:$B$55,$B10,Sheet1!L$2:L$55)'
Sheet['K11'] = '=SUMIF(Sheet1!$B$2:$B$55,$B11,Sheet1!L$2:L$55)'
Sheet['K12'] = '=SUMIF(Sheet1!$B$2:$B$55,$B12,Sheet1!L$2:L$55)'
Sheet['L2'] = '=SUMIF(Sheet1!$B$2:$B$55,$B2,Sheet1!M$2:M$55)'
Sheet['L3'] = '=SUMIF(Sheet1!$B$2:$B$55,$B3,Sheet1!M$2:M$55)'
Sheet['L4'] = '=SUMIF(Sheet1!$B$2:$B$55,$B4,Sheet1!M$2:M$55)'
Sheet['L5'] = '=SUMIF(Sheet1!$B$2:$B$55,$B5,Sheet1!M$2:M$55)'
Sheet['L6'] = '=SUMIF(Sheet1!$B$2:$B$55,$B6,Sheet1!M$2:M$55)'
Sheet['L7'] = '=SUMIF(Sheet1!$B$2:$B$55,$B7,Sheet1!M$2:M$55)'
Sheet['L8'] = '=SUMIF(Sheet1!$B$2:$B$55,$B8,Sheet1!M$2:M$55)'
Sheet['L9'] = '=SUMIF(Sheet1!$B$2:$B$55,$B9,Sheet1!M$2:M$55)'
Sheet['L10'] = '=SUMIF(Sheet1!$B$2:$B$55,$B10,Sheet1!M$2:M$55)'
Sheet['L11'] = '=SUMIF(Sheet1!$B$2:$B$55,$B11,Sheet1!M$2:M$55)'
Sheet['L12'] = '=SUMIF(Sheet1!$B$2:$B$55,$B12,Sheet1!M$2:M$55)'
Sheet['M2'] = '=SUMIF(Sheet1!$B$2:$B$55,$B2,Sheet1!N$2:N$55)'
Sheet['M3'] = '=SUMIF(Sheet1!$B$2:$B$55,$B3,Sheet1!N$2:N$55)'
Sheet['M4'] = '=SUMIF(Sheet1!$B$2:$B$55,$B4,Sheet1!N$2:N$55)'
Sheet['M5'] = '=SUMIF(Sheet1!$B$2:$B$55,$B5,Sheet1!N$2:N$55)'
Sheet['M6'] = '=SUMIF(Sheet1!$B$2:$B$55,$B6,Sheet1!N$2:N$55)'
Sheet['M7'] = '=SUMIF(Sheet1!$B$2:$B$55,$B7,Sheet1!N$2:N$55)'
Sheet['M8'] = '=SUMIF(Sheet1!$B$2:$B$55,$B8,Sheet1!N$2:N$55)'
Sheet['M9'] = '=SUMIF(Sheet1!$B$2:$B$55,$B9,Sheet1!N$2:N$55)'
Sheet['M10'] = '=SUMIF(Sheet1!$B$2:$B$55,$B10,Sheet1!N$2:N$55)'
Sheet['M11'] = '=SUMIF(Sheet1!$B$2:$B$55,$B11,Sheet1!N$2:N$55)'
Sheet['M12'] = '=SUMIF(Sheet1!$B$2:$B$55,$B12,Sheet1!N$2:N$55)'
Sheet['N2'] = '=ROUND(((L2+M2)/K2*100),2)'
Sheet['N3'] = '=ROUND(((L3+M3)/K3*100),2)'
Sheet['N4'] = '=ROUND(((L4+M4)/K4*100),2)'
Sheet['N5'] = '=ROUND(((L5+M5)/K5*100),2)'
Sheet['N6'] = '=ROUND(((L6+M6)/K6*100),2)'
Sheet['N7'] = '=ROUND(((L7+M7)/K7*100),2)'
Sheet['N8'] = '=ROUND(((L8+M8)/K8*100),2)'
Sheet['N9'] = '=ROUND(((L9+M9)/K9*100),2)'
Sheet['N10'] = '=ROUND(((L10+M10)/K10*100),2)'
Sheet['N11'] = '=ROUND(((L11+M11)/K11*100),2)'
Sheet['N12'] = '=ROUND(((L12+M12)/K12*100),2)'
Sheet['O2'] = '=ROUND((G2/H2*100),2)'
Sheet['O3'] = '=ROUND((G3/H3*100),2)'
Sheet['O4'] = '=ROUND((G4/H4*100),2)'
Sheet['O5'] = '=ROUND((G5/H5*100),2)'
Sheet['O6'] = '=ROUND((G6/H6*100),2)'
Sheet['O7'] = '=ROUND((G7/H7*100),2)'
Sheet['O8'] = '=ROUND((G8/H8*100),2)'
Sheet['O9'] = '=ROUND((G9/H9*100),2)'
Sheet['O10'] = '=ROUND((G10/H10*100),2)'
Sheet['O11'] = '=ROUND((G11/H11*100),2)'
Sheet['O12'] = '=ROUND((G12/H12*100),2)'
Sheet['P2'] = '=ROUND((E2/F2*100),2)'
Sheet['P3'] = '=ROUND((E3/F3*100),2)'
Sheet['P4'] = '=ROUND((E4/F4*100),2)'
Sheet['P5'] = '=ROUND((E5/F5*100),2)'
Sheet['P6'] = '=ROUND((E6/F6*100),2)'
Sheet['P7'] = '=ROUND((E7/F7*100),2)'
Sheet['P8'] = '=ROUND((E8/F8*100),2)'
Sheet['P9'] = '=ROUND((E9/F9*100),2)'
Sheet['P10'] = '=ROUND((E10/F10*100),2)'
Sheet['P11'] = '=ROUND((E11/F11*100),2)'
Sheet['P12'] = '=ROUND((E12/F12*100),2)'
# Sheet['J13'] = '=SUM(J2:J12)'
# Sheet['K13'] = '=SUM(K2:K12)'
# Sheet['O13'] = '=ROUND(AVERAGE(O2:O12),2)'
# Sheet['P13'] = '=ROUND(AVERAGE(P2:P12),2)'
# print("Value of the Cell 1:",Sheet['P13'].value)
Sheet1['O2'] = '=ROUND(((N2+M2)/L2*100),2)'
Sheet1['O3'] = '=ROUND(((N3+M3)/L3*100),2)'
Sheet1['O4'] = '=ROUND(((N4+M4)/L4*100),2)'
Sheet1['O5'] = '=ROUND(((N5+M5)/L5*100),2)'
Sheet1['O6'] = '=ROUND(((N6+M6)/L6*100),2)'
Sheet1['O7'] = '=ROUND(((N7+M7)/L7*100),2)'
Sheet1['O8'] = '=ROUND(((N8+M8)/L8*100),2)'
Sheet1['O9'] = '=ROUND(((N9+M9)/L9*100),2)'
Sheet1['O10'] = '=ROUND(((N10+M10)/L10*100),2)'
Sheet1['O11'] = '=ROUND(((N11+M11)/L11*100),2)'
Sheet1['O12'] = '=ROUND(((N12+M12)/L12*100),2)'
Sheet1['O13'] = '=ROUND(((N13+M13)/L13*100),2)'
Sheet1['O14'] = '=ROUND(((N14+M14)/L14*100),2)'
Sheet1['O15'] = '=ROUND(((N15+M15)/L15*100),2)'
Sheet1['O16'] = '=ROUND(((N16+M16)/L16*100),2)'
Sheet1['O17'] = '=ROUND(((N17+M17)/L17*100),2)'
Sheet1['O18'] = '=ROUND(((N18+M18)/L18*100),2)'
Sheet1['O19'] = '=ROUND(((N19+M19)/L19*100),2)'
Sheet1['O20'] = '=ROUND(((N20+M20)/L20*100),2)'
Sheet1['O21'] = '=ROUND(((N21+M21)/L21*100),2)'
Sheet1['O22'] = '=ROUND(((N22+M22)/L22*100),2)'
Sheet1['O23'] = '=ROUND(((N23+M23)/L23*100),2)'
Sheet1['O24'] = '=ROUND(((N24+M24)/L24*100),2)'
Sheet1['O25'] = '=ROUND(((N25+M25)/L25*100),2)'
Sheet1['O26'] = '=ROUND(((N26+M26)/L26*100),2)'
Sheet1['O27'] = '=ROUND(((N27+M27)/L27*100),2)'
Sheet1['O28'] = '=ROUND(((N28+M28)/L28*100),2)'
Sheet1['O29'] = '=ROUND(((N29+M29)/L29*100),2)'
Sheet1['O30'] = '=ROUND(((N30+M30)/L30*100),2)'
Sheet1['O31'] = '=ROUND(((N31+M31)/L31*100),2)'
Sheet1['O32'] = '=ROUND(((N32+M32)/L32*100),2)'
Sheet1['O33'] = '=ROUND(((N33+M33)/L33*100),2)'
Sheet1['O34'] = '=ROUND(((N34+M34)/L34*100),2)'
Sheet1['O35'] = '=ROUND(((N35+M35)/L35*100),2)'
Sheet1['O36'] = '=ROUND(((N36+M36)/L36*100),2)'
Sheet1['O37'] = '=ROUND(((N37+M37)/L37*100),2)'
Sheet1['O38'] = '=ROUND(((N38+M38)/L38*100),2)'
Sheet1['O39'] = '=ROUND(((N39+M39)/L39*100),2)'
Sheet1['O40'] = '=ROUND(((N40+M40)/L40*100),2)'
Sheet1['O41'] = '=ROUND(((N41+M41)/L41*100),2)'
Sheet1['O42'] = '=ROUND(((N42+M42)/L42*100),2)'
Sheet1['O43'] = '=ROUND(((N43+M43)/L43*100),2)'
Sheet1['O44'] = '=ROUND(((N44+M44)/L44*100),2)'
Sheet1['O45'] = '=ROUND(((N45+M45)/L45*100),2)'
Sheet1['O46'] = '=ROUND(((N46+M46)/L46*100),2)'
Sheet1['O47'] = '=ROUND(((N47+M47)/L47*100),2)'
Sheet1['O48'] = '=ROUND(((N48+M48)/L48*100),2)'
Sheet1['O49'] = '=ROUND(((N49+M49)/L49*100),2)'
Sheet1['O50'] = '=ROUND(((N50+M50)/L50*100),2)'
Sheet1['O51'] = '=ROUND(((N51+M51)/L51*100),2)'
Sheet1['O52'] = '=ROUND(((N52+M52)/L52*100),2)'
Sheet1['O53'] = '=ROUND(((N53+M53)/L53*100),2)'
Sheet1['O54'] = '=ROUND(((N54+M54)/L54*100),2)'
Sheet1['O55'] = '=ROUND(((N55+M55)/L55*100),2)'

Sheet1['P2'] = '=ROUND((F2/G2*100),2)'
Sheet1['P3'] = '=ROUND((F3/G3*100),2)'
Sheet1['P4'] = '=ROUND((F4/G4*100),2)'
Sheet1['P5'] = '=ROUND((F5/G5*100),2)'
Sheet1['P6'] = '=ROUND((F6/G6*100),2)'
Sheet1['P7'] = '=ROUND((F7/G7*100),2)'
Sheet1['P8'] = '=ROUND((F8/G8*100),2)'
Sheet1['P9'] = '=ROUND((F9/G9*100),2)'
Sheet1['P10'] = '=ROUND((F10/G10*100),2)'
Sheet1['P11'] = '=ROUND((F11/G11*100),2)'
Sheet1['P12'] = '=ROUND((F12/G12*100),2)'
Sheet1['P13'] = '=ROUND((F13/G13*100),2)'
Sheet1['P14'] = '=ROUND((F14/G14*100),2)'
Sheet1['P15'] = '=ROUND((F15/G15*100),2)'
Sheet1['P16'] = '=ROUND((F16/G16*100),2)'
Sheet1['P17'] = '=ROUND((F17/G17*100),2)'
Sheet1['P18'] = '=ROUND((F18/G18*100),2)'
Sheet1['P19'] = '=ROUND((F19/G19*100),2)'
Sheet1['P20'] = '=ROUND((F20/G20*100),2)'
Sheet1['P21'] = '=ROUND((F21/G21*100),2)'
Sheet1['P22'] = '=ROUND((F22/G22*100),2)'
Sheet1['P23'] = '=ROUND((F23/G23*100),2)'
Sheet1['P24'] = '=ROUND((F24/G24*100),2)'
Sheet1['P25'] = '=ROUND((F25/G25*100),2)'
Sheet1['P26'] = '=ROUND((F26/G26*100),2)'
Sheet1['P27'] = '=ROUND((F27/G27*100),2)'
Sheet1['P28'] = '=ROUND((F28/G28*100),2)'
Sheet1['P29'] = '=ROUND((F29/G29*100),2)'
Sheet1['P30'] = '=ROUND((F30/G30*100),2)'
Sheet1['P31'] = '=ROUND((F31/G31*100),2)'
Sheet1['P32'] = '=ROUND((F32/G32*100),2)'
Sheet1['P33'] = '=ROUND((F33/G33*100),2)'
Sheet1['P34'] = '=ROUND((F34/G34*100),2)'
Sheet1['P35'] = '=ROUND((F35/G35*100),2)'
Sheet1['P36'] = '=ROUND((F36/G36*100),2)'
Sheet1['P37'] = '=ROUND((F37/G37*100),2)'
Sheet1['P38'] = '=ROUND((F38/G38*100),2)'
Sheet1['P39'] = '=ROUND((F39/G39*100),2)'
Sheet1['P40'] = '=ROUND((F40/G40*100),2)'
Sheet1['P41'] = '=ROUND((F41/G41*100),2)'
Sheet1['P42'] = '=ROUND((F42/G42*100),2)'
Sheet1['P43'] = '=ROUND((F43/G43*100),2)'
Sheet1['P44'] = '=ROUND((F44/G44*100),2)'
Sheet1['P45'] = '=ROUND((F45/G45*100),2)'
Sheet1['P46'] = '=ROUND((F46/G46*100),2)'
Sheet1['P47'] = '=ROUND((F47/G47*100),2)'
Sheet1['P48'] = '=ROUND((F48/G48*100),2)'
Sheet1['P49'] = '=ROUND((F49/G49*100),2)'
Sheet1['P50'] = '=ROUND((F50/G50*100),2)'
Sheet1['P51'] = '=ROUND((F51/G51*100),2)'
Sheet1['P52'] = '=ROUND((F52/G52*100),2)'
Sheet1['P53'] = '=ROUND((F53/G53*100),2)'
Sheet1['P54'] = '=ROUND((F54/G54*100),2)'
Sheet1['P55'] = '=ROUND((F55/G55*100),2)'
Sheet1['Q2'] = '=ROUND((H2/I2*100),2)'
Sheet1['Q3'] = '=ROUND((H3/I3*100),2)'
Sheet1['Q4'] = '=ROUND((H4/I4*100),2)'
Sheet1['Q5'] = '=ROUND((H5/I5*100),2)'
Sheet1['Q6'] = '=ROUND((H6/I6*100),2)'
Sheet1['Q7'] = '=ROUND((H7/I7*100),2)'
Sheet1['Q8'] = '=ROUND((H8/I8*100),2)'
Sheet1['Q9'] = '=ROUND((H9/I9*100),2)'
Sheet1['Q10'] = '=ROUND((H10/I10*100),2)'
Sheet1['Q11'] = '=ROUND((H11/I11*100),2)'
Sheet1['Q12'] = '=ROUND((H12/I12*100),2)'
Sheet1['Q13'] = '=ROUND((H13/I13*100),2)'
Sheet1['Q14'] = '=ROUND((H14/I14*100),2)'
Sheet1['Q15'] = '=ROUND((H15/I15*100),2)'
Sheet1['Q16'] = '=ROUND((H16/I16*100),2)'
Sheet1['Q17'] = '=ROUND((H17/I17*100),2)'
Sheet1['Q18'] = '=ROUND((H18/I18*100),2)'
Sheet1['Q19'] = '=ROUND((H19/I19*100),2)'
Sheet1['Q20'] = '=ROUND((H20/I20*100),2)'
Sheet1['Q21'] = '=ROUND((H21/I21*100),2)'
Sheet1['Q22'] = '=ROUND((H22/I22*100),2)'
Sheet1['Q23'] = '=ROUND((H23/I23*100),2)'
Sheet1['Q24'] = '=ROUND((H24/I24*100),2)'
Sheet1['Q25'] = '=ROUND((H25/I25*100),2)'
Sheet1['Q26'] = '=ROUND((H26/I26*100),2)'
Sheet1['Q27'] = '=ROUND((H27/I27*100),2)'
Sheet1['Q28'] = '=ROUND((H28/I28*100),2)'
Sheet1['Q29'] = '=ROUND((H29/I29*100),2)'
Sheet1['Q30'] = '=ROUND((H30/I30*100),2)'
Sheet1['Q31'] = '=ROUND((H31/I31*100),2)'
Sheet1['Q32'] = '=ROUND((H32/I32*100),2)'
Sheet1['Q33'] = '=ROUND((H33/I33*100),2)'
Sheet1['Q34'] = '=ROUND((H34/I34*100),2)'
Sheet1['Q35'] = '=ROUND((H35/I35*100),2)'
Sheet1['Q36'] = '=ROUND((H36/I36*100),2)'
Sheet1['Q37'] = '=ROUND((H37/I37*100),2)'
Sheet1['Q38'] = '=ROUND((H38/I38*100),2)'
Sheet1['Q39'] = '=ROUND((H39/I39*100),2)'
Sheet1['Q40'] = '=ROUND((H40/I40*100),2)'
Sheet1['Q41'] = '=ROUND((H41/I41*100),2)'
Sheet1['Q42'] = '=ROUND((H42/I42*100),2)'
Sheet1['Q43'] = '=ROUND((H43/I43*100),2)'
Sheet1['Q44'] = '=ROUND((H44/I44*100),2)'
Sheet1['Q45'] = '=ROUND((H45/I45*100),2)'
Sheet1['Q46'] = '=ROUND((H46/I46*100),2)'
Sheet1['Q47'] = '=ROUND((H47/I47*100),2)'
Sheet1['Q48'] = '=ROUND((H48/I48*100),2)'
Sheet1['Q49'] = '=ROUND((H49/I49*100),2)'
Sheet1['Q50'] = '=ROUND((H50/I50*100),2)'
Sheet1['Q51'] = '=ROUND((H51/I51*100),2)'
Sheet1['Q52'] = '=ROUND((H52/I52*100),2)'
Sheet1['Q53'] = '=ROUND((H53/I53*100),2)'
Sheet1['Q54'] = '=ROUND((H54/I54*100),2)'
Sheet1['Q55'] = '=ROUND((H55/I55*100),2)'


wb.save(xlsx_file_name)


today = date.today()
date_str = datetime.now().strftime('%Y-%m-%d')
email_address = "fabtrakr@ideplexports.com"
email_password = "%iDflo@777%"
xlsx_file_name =  'file-'+date_str+'.xlsx'
msg = EmailMessage()
msg['Subject'] = "Overview Report : " +date_str
msg['From'] = email_address
msg['To'] = "gurumurthyrudresh@gmail.com"
# msg['To'] = "gurumurthyrudresh@gmail.com,madhukar@indian-designs.com"
msg['Cc'] = "gurumurthy@indian-designs.com"
msg.set_content("""Dear Team,

Please find the below Overview report for """+str(date_str)+""".

Regards
Team Idflo
This is an Auto Generated Email, Please don't reply.""")

with open(xlsx_file_name,"rb") as f:
    file_data=f.read()
    file_name=f.name
    print(file_name)
    msg.add_attachment(file_data,maintype="application",subtype="xlsx",filename=file_name)
    
# send email
with smtplib.SMTP_SSL('smtp3.netcore.co.in', 465) as smtp:
    smtp.login(email_address, email_password)
    smtp.send_message(msg)
