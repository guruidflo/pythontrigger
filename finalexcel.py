import openpyxl
from datetime import datetime
from datetime import date


date_today  = date.today() 
date_str = datetime.now().strftime('%Y-%m-%d')
xlsx_file_name =  'file-'+date_str+'.xlsx'
xlsx_file_combined =  'combined-'+date_str+'.xlsx'

source_wb = openpyxl.load_workbook(xlsx_file_combined)
target_wb = openpyxl.load_workbook(xlsx_file_name)

# Select the sheet from the source workbook
source_sheet = source_wb['Sheet1']

# Select or Create the sheet in the target workbook
if 'Sheet2' in target_wb.sheetnames:
    target_sheet = target_wb['Sheet2']
    target_wb.remove(target_sheet)
    target_wb.create_sheet('Sheet2')
else:
    target_sheet = target_wb.create_sheet('Sheet2')

# Copy the data from the source sheet to the target sheet
for row in source_sheet.iter_rows():
    for cell in row:
        target_sheet[cell.coordinate].value = cell.value

# Save the target workbook
target_wb.save(xlsx_file_name)
wb = openpyxl.load_workbook(xlsx_file_name)
sheet = wb['Sheet1']

# Insert the new columns
# sheet.insert_cols(6, 4)

wb.save(xlsx_file_name)