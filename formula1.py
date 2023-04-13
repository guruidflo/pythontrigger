import openpyxl
from datetime import datetime
from datetime import timedelta
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import Alignment,Font,PatternFill

date_str = datetime.now().strftime('%Y-%m-%d')
json_file_name = 'file-'+date_str+'.json'
xlsx_file_name =  'file-'+date_str+'.xlsx'

wb = openpyxl.load_workbook(xlsx_file_name)
Sheet = wb.get_sheet_by_name('Summary')
Sheet1 = wb.get_sheet_by_name('Sheet1')

columns = ['p']

last_row = Sheet1.max_row

for column in columns:
    for row in range(2, last_row + 1):
        Sheet1[column + str(row)].value = '=ROUND(((N2+O2)/M2*100),2)'

wb.save(xlsx_file_name)
