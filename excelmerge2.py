import openpyxl
from datetime import datetime
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import Alignment,Font,PatternFill
from openpyxl.styles import Border, Side


date_today  = date.today() 
date_str = datetime.now().strftime('%Y-%m-%d')
xlsx_file_name =  'file-'+date_str+'.xlsx'


# Open the workbook
wb = openpyxl.load_workbook(xlsx_file_name)
Sheet1 = wb.get_sheet_by_name('Sheet1')

# Get the Sheet1 and Sheet2
sheet1 = wb['Sheet1']
sheet2 = wb['Sheet2']

num_rows_sheet1 = sheet1.max_row
num_rows_sheet2 = sheet2.max_row

# Copy values from FGHI columns of Sheet2 to FGHI columns of Sheet1 if the C column values match
for i in range(2, num_rows_sheet2 + 1):
    for j in range(2, num_rows_sheet1 + 1):
        if sheet2.cell(row=i, column=3).value == sheet1.cell(row=j, column=3).value:
            sheet1.cell(row=j, column=6).value = sheet2.cell(row=i, column=6).value
            sheet1.cell(row=j, column=7).value = sheet2.cell(row=i, column=7).value
            sheet1.cell(row=j, column=8).value = sheet2.cell(row=i, column=8).value
            sheet1.cell(row=j, column=9).value = sheet2.cell(row=i, column=9).value


for row in Sheet1.iter_rows():
    for cell in row:
        cell.alignment = openpyxl.styles.Alignment(horizontal='center')

thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))

for row in Sheet1.iter_rows():
    for cell in row:
        if cell.value:
            cell.border = thin_border

        


wb.save(xlsx_file_name)
