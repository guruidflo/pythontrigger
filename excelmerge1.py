import openpyxl

# Open the first workbook
wb1 = openpyxl.load_workbook('file-2023-01-16.xlsx')

# Open the second workbook
wb2 = openpyxl.load_workbook('Output.xlsx')

# Iterate through each sheet in the second workbook
for sheet in wb2.sheetnames:
    # Copy the sheet to the first workbook
    ws = wb2[sheet]
    wb1.create_sheet(title=sheet)
    wb1[sheet].append(ws)

# Save the merged workbook
wb1.save('merged_file.xlsx')
