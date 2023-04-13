import openpyxl
from datetime import datetime
from datetime import timedelta
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import Alignment,Font,PatternFill
from openpyxl.styles import Border, Side


date_str = datetime.now().strftime('%Y-%m-%d')
json_file_name = 'file-'+date_str+'.json'
xlsx_file_name =  'file-'+date_str+'.xlsx'

wb = openpyxl.load_workbook(xlsx_file_name)
sheet_name = "Summary"
wb.create_sheet(title=sheet_name)

wb.save(xlsx_file_name)


wb = openpyxl.load_workbook(xlsx_file_name)
Sheet = wb.get_sheet_by_name('Summary')
range = Sheet['A3':'Q16']

Sheet1 = wb.get_sheet_by_name('Sheet1')
Sheet1.insert_cols(6, 4)


Sheet['A3'] = 'Sl No.'
Sheet['B3'] = 'location'
Sheet['C3'] = 'Target'
Sheet['A4'] = '1'
Sheet['A5'] = '2'
Sheet['A6'] = '3'
Sheet['A7'] = '4'
Sheet['A8'] = '5'
Sheet['A9'] = '6'
Sheet['A10'] = '7'
Sheet['A11'] = '8'
Sheet['A12'] = '9'
Sheet['A13'] = '10'
Sheet['A14'] = '11'
Sheet['A15'] = '12'

Sheet['B4'] = 'IDD1'
Sheet['B5'] = 'IDU1'
Sheet['B6'] = 'IDU3'
Sheet['B7'] = 'IDU4'
Sheet['B8'] = 'IDU5'
Sheet['B9'] = 'IDU6'
Sheet['B10'] = 'IDU8'
Sheet['B11'] = 'IDU9P1'
Sheet['B12'] = 'NJU1P1'
Sheet['B13'] = 'NJU1P2'
Sheet['B14'] = 'NJU2'
Sheet['B15'] = 'NJU3'

Sheet['D3'] = 'Output'
Sheet['E3'] = 'Sewing Sam Produced'
Sheet['F3'] = 'Sewing Available Minutes'
Sheet['G3'] = 'Global Sam Produced'
Sheet['H3'] = 'Global Available Minutes'
Sheet['I3'] = 'Total Work Station'
Sheet['J3'] = 'Total Machines'
Sheet['K3'] = 'End Table pass'
Sheet['L3'] = 'Defects'
Sheet['M3'] = 'Rejection'
Sheet['N3'] = 'DHU%'
Sheet['P3'] = 'SEWING EFFICIENCY %'
Sheet['O3'] = 'GLOBAL EFFICIENCY %'

Sheet1['B1'] = 'Unit'
Sheet1['C1'] = 'BatchID'
Sheet1['D1'] = 'Batch'
Sheet1['E1'] = 'Target'
Sheet1['F1'] = 'OC-Number'	
Sheet1['G1'] = 'ProductCode'	
Sheet1['H1'] = 'ProductDescription'	
Sheet1['I1'] = 'BuyerName'
Sheet1['J1'] = 'Output'
Sheet1['K1'] = 'Sewing Sam Produced'
Sheet1['L1'] = 'Sewing Available Minutes'
Sheet1['M1'] = 'Global Sam Produced'
Sheet1['N1'] = 'Global Available Minutes'
Sheet1['O1'] = 'Total Work Station'
Sheet1['P1'] = 'Total Machines'
Sheet1['Q1'] = 'End Table pass'
Sheet1['R1'] = 'Defects'
Sheet1['S1'] = 'Rejection'
Sheet1['T1'] = 'DHU%'
Sheet1['U1'] = 'SEWING EFFICIENCY %'
Sheet1['V1'] = 'GLOBAL EFFICIENCY %'

Sheet['C4'] = '=SUMIF(Sheet1!$B$2:$B$100,$B4,Sheet1!E$2:E$100)'
Sheet['C5'] = '=SUMIF(Sheet1!$B$2:$B$100,$B5,Sheet1!E$2:E$100)'
Sheet['C6'] = '=SUMIF(Sheet1!$B$2:$B$100,$B6,Sheet1!E$2:E$100)'
Sheet['C7'] = '=SUMIF(Sheet1!$B$2:$B$100,$B7,Sheet1!E$2:E$100)'
Sheet['C8'] = '=SUMIF(Sheet1!$B$2:$B$100,$B8,Sheet1!E$2:E$100)'
Sheet['C9'] = '=SUMIF(Sheet1!$B$2:$B$100,$B9,Sheet1!E$2:E$100)'
Sheet['C10'] = '=SUMIF(Sheet1!$B$2:$B$100,$B10,Sheet1!E$2:E$100)'
Sheet['C11'] = '=SUMIF(Sheet1!$B$2:$B$100,$B11,Sheet1!E$2:E$100)'
Sheet['C12'] = '=SUMIF(Sheet1!$B$2:$B$100,$B12,Sheet1!E$2:E$100)'
Sheet['C13'] = '=SUMIF(Sheet1!$B$2:$B$100,$B13,Sheet1!E$2:E$100)'
Sheet['C14'] = '=SUMIF(Sheet1!$B$2:$B$100,$B14,Sheet1!E$2:E$100)'
Sheet['C15'] = '=SUMIF(Sheet1!$B$2:$B$100,$B15,Sheet1!E$2:E$100)'

Sheet['D4'] = '=SUMIF(Sheet1!$B$2:$B$100,$B4,Sheet1!J$2:J$100)'
Sheet['D5'] = '=SUMIF(Sheet1!$B$2:$B$100,$B5,Sheet1!J$2:J$100)'
Sheet['D6'] = '=SUMIF(Sheet1!$B$2:$B$100,$B6,Sheet1!J$2:J$100)'
Sheet['D7'] = '=SUMIF(Sheet1!$B$2:$B$100,$B7,Sheet1!J$2:J$100)'
Sheet['D8'] = '=SUMIF(Sheet1!$B$2:$B$100,$B8,Sheet1!J$2:J$100)'
Sheet['D9'] = '=SUMIF(Sheet1!$B$2:$B$100,$B9,Sheet1!J$2:J$100)'
Sheet['D10'] = '=SUMIF(Sheet1!$B$2:$B$100,$B10,Sheet1!J$2:J$100)'
Sheet['D11'] = '=SUMIF(Sheet1!$B$2:$B$100,$B11,Sheet1!J$2:J$100)'
Sheet['D12'] = '=SUMIF(Sheet1!$B$2:$B$100,$B12,Sheet1!J$2:J$100)'
Sheet['D13'] = '=SUMIF(Sheet1!$B$2:$B$100,$B13,Sheet1!J$2:J$100)'
Sheet['D14'] = '=SUMIF(Sheet1!$B$2:$B$100,$B14,Sheet1!J$2:J$100)'
Sheet['D15'] = '=SUMIF(Sheet1!$B$2:$B$100,$B15,Sheet1!J$2:J$100)'

Sheet['E4'] = '=SUMIF(Sheet1!$B$2:$B$100,$B4,Sheet1!K$2:K$100)'
Sheet['E5'] = '=SUMIF(Sheet1!$B$2:$B$100,$B5,Sheet1!K$2:K$100)'
Sheet['E6'] = '=SUMIF(Sheet1!$B$2:$B$100,$B6,Sheet1!K$2:K$100)'
Sheet['E7'] = '=SUMIF(Sheet1!$B$2:$B$100,$B7,Sheet1!K$2:K$100)'
Sheet['E8'] = '=SUMIF(Sheet1!$B$2:$B$100,$B8,Sheet1!K$2:K$100)'
Sheet['E9'] = '=SUMIF(Sheet1!$B$2:$B$100,$B9,Sheet1!K$2:K$100)'
Sheet['E10'] = '=SUMIF(Sheet1!$B$2:$B$100,$B10,Sheet1!K$2:K$100)'
Sheet['E11'] = '=SUMIF(Sheet1!$B$2:$B$100,$B11,Sheet1!K$2:K$100)'
Sheet['E12'] = '=SUMIF(Sheet1!$B$2:$B$100,$B12,Sheet1!K$2:K$100)'
Sheet['E13'] = '=SUMIF(Sheet1!$B$2:$B$100,$B13,Sheet1!K$2:K$100)'
Sheet['E14'] = '=SUMIF(Sheet1!$B$2:$B$100,$B14,Sheet1!K$2:K$100)'
Sheet['E15'] = '=SUMIF(Sheet1!$B$2:$B$100,$B15,Sheet1!K$2:K$100)'

Sheet['F4'] = '=SUMIF(Sheet1!$B$2:$B$100,$B4,Sheet1!L$2:L$100)'
Sheet['F5'] = '=SUMIF(Sheet1!$B$2:$B$100,$B5,Sheet1!L$2:L$100)'
Sheet['F6'] = '=SUMIF(Sheet1!$B$2:$B$100,$B6,Sheet1!L$2:L$100)'
Sheet['F7'] = '=SUMIF(Sheet1!$B$2:$B$100,$B7,Sheet1!L$2:L$100)'
Sheet['F8'] = '=SUMIF(Sheet1!$B$2:$B$100,$B8,Sheet1!L$2:L$100)'
Sheet['F9'] = '=SUMIF(Sheet1!$B$2:$B$100,$B9,Sheet1!L$2:L$100)'
Sheet['F10'] = '=SUMIF(Sheet1!$B$2:$B$100,$B10,Sheet1!L$2:L$100)'
Sheet['F11'] = '=SUMIF(Sheet1!$B$2:$B$100,$B11,Sheet1!L$2:L$100)'
Sheet['F12'] = '=SUMIF(Sheet1!$B$2:$B$100,$B12,Sheet1!L$2:L$100)'
Sheet['F13'] = '=SUMIF(Sheet1!$B$2:$B$100,$B13,Sheet1!L$2:L$100)'
Sheet['F14'] = '=SUMIF(Sheet1!$B$2:$B$100,$B14,Sheet1!L$2:L$100)'
Sheet['F15'] = '=SUMIF(Sheet1!$B$2:$B$100,$B15,Sheet1!L$2:L$100)'

Sheet['G4'] = '=SUMIF(Sheet1!$B$2:$B$100,$B4,Sheet1!M$2:M$100)'
Sheet['G5'] = '=SUMIF(Sheet1!$B$2:$B$100,$B5,Sheet1!M$2:M$100)'
Sheet['G6'] = '=SUMIF(Sheet1!$B$2:$B$100,$B6,Sheet1!M$2:M$100)'
Sheet['G7'] = '=SUMIF(Sheet1!$B$2:$B$100,$B7,Sheet1!M$2:M$100)'
Sheet['G8'] = '=SUMIF(Sheet1!$B$2:$B$100,$B8,Sheet1!M$2:M$100)'
Sheet['G9'] = '=SUMIF(Sheet1!$B$2:$B$100,$B9,Sheet1!M$2:M$100)'
Sheet['G10'] = '=SUMIF(Sheet1!$B$2:$B$100,$B10,Sheet1!M$2:M$100)'
Sheet['G11'] = '=SUMIF(Sheet1!$B$2:$B$100,$B11,Sheet1!M$2:M$100)'
Sheet['G12'] = '=SUMIF(Sheet1!$B$2:$B$100,$B12,Sheet1!M$2:M$100)'
Sheet['G13'] = '=SUMIF(Sheet1!$B$2:$B$100,$B13,Sheet1!M$2:M$100)'
Sheet['G14'] = '=SUMIF(Sheet1!$B$2:$B$100,$B14,Sheet1!M$2:M$100)'
Sheet['G15'] = '=SUMIF(Sheet1!$B$2:$B$100,$B15,Sheet1!M$2:M$100)'

Sheet['H4'] = '=SUMIF(Sheet1!$B$2:$B$100,$B4,Sheet1!N$2:N$100)'
Sheet['H5'] = '=SUMIF(Sheet1!$B$2:$B$100,$B5,Sheet1!N$2:N$100)'
Sheet['H6'] = '=SUMIF(Sheet1!$B$2:$B$100,$B6,Sheet1!N$2:N$100)'
Sheet['H7'] = '=SUMIF(Sheet1!$B$2:$B$100,$B7,Sheet1!N$2:N$100)'
Sheet['H8'] = '=SUMIF(Sheet1!$B$2:$B$100,$B8,Sheet1!N$2:N$100)'
Sheet['H9'] = '=SUMIF(Sheet1!$B$2:$B$100,$B9,Sheet1!N$2:N$100)'
Sheet['H10'] = '=SUMIF(Sheet1!$B$2:$B$100,$B10,Sheet1!N$2:N$100)'
Sheet['H11'] = '=SUMIF(Sheet1!$B$2:$B$100,$B11,Sheet1!N$2:N$100)'
Sheet['H12'] = '=SUMIF(Sheet1!$B$2:$B$100,$B12,Sheet1!N$2:N$100)'
Sheet['H13'] = '=SUMIF(Sheet1!$B$2:$B$100,$B13,Sheet1!N$2:N$100)'
Sheet['H14'] = '=SUMIF(Sheet1!$B$2:$B$100,$B14,Sheet1!N$2:N$100)'
Sheet['H15'] = '=SUMIF(Sheet1!$B$2:$B$100,$B15,Sheet1!N$2:N$100)'

Sheet['I4'] = '=SUMIF(Sheet1!$B$2:$B$100,$B4,Sheet1!O$2:O$100)'
Sheet['I5'] = '=SUMIF(Sheet1!$B$2:$B$100,$B5,Sheet1!O$2:O$100)'
Sheet['I6'] = '=SUMIF(Sheet1!$B$2:$B$100,$B6,Sheet1!O$2:O$100)'
Sheet['I7'] = '=SUMIF(Sheet1!$B$2:$B$100,$B7,Sheet1!O$2:O$100)'
Sheet['I8'] = '=SUMIF(Sheet1!$B$2:$B$100,$B8,Sheet1!O$2:O$100)'
Sheet['I9'] = '=SUMIF(Sheet1!$B$2:$B$100,$B9,Sheet1!O$2:O$100)'
Sheet['I10'] = '=SUMIF(Sheet1!$B$2:$B$100,$B10,Sheet1!O$2:O$100)'
Sheet['I11'] = '=SUMIF(Sheet1!$B$2:$B$100,$B11,Sheet1!O$2:O$100)'
Sheet['I12'] = '=SUMIF(Sheet1!$B$2:$B$100,$B12,Sheet1!O$2:O$100)'
Sheet['I13'] = '=SUMIF(Sheet1!$B$2:$B$100,$B13,Sheet1!O$2:O$100)'
Sheet['I14'] = '=SUMIF(Sheet1!$B$2:$B$100,$B14,Sheet1!O$2:O$100)'
Sheet['I15'] = '=SUMIF(Sheet1!$B$2:$B$100,$B15,Sheet1!O$2:O$100)'

Sheet['J4'] = '=SUMIF(Sheet1!$B$2:$B$100,$B4,Sheet1!P$2:P$100)'
Sheet['J5'] = '=SUMIF(Sheet1!$B$2:$B$100,$B5,Sheet1!P$2:P$100)'
Sheet['J6'] = '=SUMIF(Sheet1!$B$2:$B$100,$B6,Sheet1!P$2:P$100)'
Sheet['J7'] = '=SUMIF(Sheet1!$B$2:$B$100,$B7,Sheet1!P$2:P$100)'
Sheet['J8'] = '=SUMIF(Sheet1!$B$2:$B$100,$B8,Sheet1!P$2:P$100)'
Sheet['J9'] = '=SUMIF(Sheet1!$B$2:$B$100,$B9,Sheet1!P$2:P$100)'
Sheet['J10'] = '=SUMIF(Sheet1!$B$2:$B$100,$B10,Sheet1!P$2:P$100)'
Sheet['J11'] = '=SUMIF(Sheet1!$B$2:$B$100,$B11,Sheet1!P$2:P$100)'
Sheet['J12'] = '=SUMIF(Sheet1!$B$2:$B$100,$B12,Sheet1!P$2:P$100)'
Sheet['J13'] = '=SUMIF(Sheet1!$B$2:$B$100,$B13,Sheet1!P$2:P$100)'
Sheet['J14'] = '=SUMIF(Sheet1!$B$2:$B$100,$B14,Sheet1!P$2:P$100)'
Sheet['J15'] = '=SUMIF(Sheet1!$B$2:$B$100,$B15,Sheet1!P$2:P$100)'

Sheet['K4'] = '=SUMIF(Sheet1!$B$2:$B$100,$B4,Sheet1!Q$2:Q$100)'
Sheet['K5'] = '=SUMIF(Sheet1!$B$2:$B$100,$B5,Sheet1!Q$2:Q$100)'
Sheet['K6'] = '=SUMIF(Sheet1!$B$2:$B$100,$B6,Sheet1!Q$2:Q$100)'
Sheet['K7'] = '=SUMIF(Sheet1!$B$2:$B$100,$B7,Sheet1!Q$2:Q$100)'
Sheet['K8'] = '=SUMIF(Sheet1!$B$2:$B$100,$B8,Sheet1!Q$2:Q$100)'
Sheet['K9'] = '=SUMIF(Sheet1!$B$2:$B$100,$B9,Sheet1!Q$2:Q$100)'
Sheet['K10'] = '=SUMIF(Sheet1!$B$2:$B$100,$B10,Sheet1!Q$2:Q$100)'
Sheet['K11'] = '=SUMIF(Sheet1!$B$2:$B$100,$B11,Sheet1!Q$2:Q$100)'
Sheet['K12'] = '=SUMIF(Sheet1!$B$2:$B$100,$B12,Sheet1!Q$2:Q$100)'
Sheet['K13'] = '=SUMIF(Sheet1!$B$2:$B$100,$B13,Sheet1!Q$2:Q$100)'
Sheet['K14'] = '=SUMIF(Sheet1!$B$2:$B$100,$B14,Sheet1!Q$2:Q$100)'
Sheet['K15'] = '=SUMIF(Sheet1!$B$2:$B$100,$B15,Sheet1!Q$2:Q$100)'

Sheet['L4'] = '=SUMIF(Sheet1!$B$2:$B$100,$B4,Sheet1!R$2:R$100)'
Sheet['L5'] = '=SUMIF(Sheet1!$B$2:$B$100,$B5,Sheet1!R$2:R$100)'
Sheet['L6'] = '=SUMIF(Sheet1!$B$2:$B$100,$B6,Sheet1!R$2:R$100)'
Sheet['L7'] = '=SUMIF(Sheet1!$B$2:$B$100,$B7,Sheet1!R$2:R$100)'
Sheet['L8'] = '=SUMIF(Sheet1!$B$2:$B$100,$B8,Sheet1!R$2:R$100)'
Sheet['L9'] = '=SUMIF(Sheet1!$B$2:$B$100,$B9,Sheet1!R$2:R$100)'
Sheet['L10'] = '=SUMIF(Sheet1!$B$2:$B$100,$B10,Sheet1!R$2:R$100)'
Sheet['L11'] = '=SUMIF(Sheet1!$B$2:$B$100,$B11,Sheet1!R$2:R$100)'
Sheet['L12'] = '=SUMIF(Sheet1!$B$2:$B$100,$B12,Sheet1!R$2:R$100)'
Sheet['L13'] = '=SUMIF(Sheet1!$B$2:$B$100,$B13,Sheet1!R$2:R$100)'
Sheet['L14'] = '=SUMIF(Sheet1!$B$2:$B$100,$B14,Sheet1!R$2:R$100)'
Sheet['L15'] = '=SUMIF(Sheet1!$B$2:$B$100,$B15,Sheet1!R$2:R$100)'

Sheet['M4'] = '=SUMIF(Sheet1!$B$2:$B$100,$B4,Sheet1!S$2:S$100)'
Sheet['M5'] = '=SUMIF(Sheet1!$B$2:$B$100,$B5,Sheet1!S$2:S$100)'
Sheet['M6'] = '=SUMIF(Sheet1!$B$2:$B$100,$B6,Sheet1!S$2:S$100)'
Sheet['M7'] = '=SUMIF(Sheet1!$B$2:$B$100,$B7,Sheet1!S$2:S$100)'
Sheet['M8'] = '=SUMIF(Sheet1!$B$2:$B$100,$B8,Sheet1!S$2:S$100)'
Sheet['M9'] = '=SUMIF(Sheet1!$B$2:$B$100,$B9,Sheet1!S$2:S$100)'
Sheet['M10'] = '=SUMIF(Sheet1!$B$2:$B$100,$B10,Sheet1!S$2:S$100)'
Sheet['M11'] = '=SUMIF(Sheet1!$B$2:$B$100,$B11,Sheet1!S$2:S$100)'
Sheet['M12'] = '=SUMIF(Sheet1!$B$2:$B$100,$B12,Sheet1!S$2:S$100)'
Sheet['M13'] = '=SUMIF(Sheet1!$B$2:$B$100,$B13,Sheet1!S$2:S$100)'
Sheet['M14'] = '=SUMIF(Sheet1!$B$2:$B$100,$B14,Sheet1!S$2:S$100)'
Sheet['M15'] = '=SUMIF(Sheet1!$B$2:$B$100,$B15,Sheet1!S$2:S$100)'


Sheet['N4'] = '=ROUND(((L4+M4)/K4*100),2)'
Sheet['N5'] = '=ROUND(((L5+M5)/K5*100),2)'
Sheet['N6'] = '=ROUND(((L6+M6)/K6*100),2)'
Sheet['N7'] = '=ROUND(((L7+M7)/K7*100),2)'
Sheet['N8'] = '=ROUND(((L8+M8)/K8*100),2)'
Sheet['N9'] = '=ROUND(((L9+M9)/K9*100),2)'
Sheet['N10'] = '=ROUND(((L10+M10)/K10*100),2)'
Sheet['N11'] = '=ROUND(((L11+M11)/K11*100),2)'
Sheet['N12'] = '=ROUND(((L12+M12)/K12*100),2)'
Sheet['N13'] = '=ROUND(((L13+M13)/K13*100),2)'
Sheet['N14'] = '=ROUND(((L14+M14)/K14*100),2)'
Sheet['N15'] = '=ROUND(((L15+M15)/K15*100),2)'


Sheet['O4'] = '=ROUND((G4/H4*100),2)'
Sheet['O5'] = '=ROUND((G5/H5*100),2)'
Sheet['O6'] = '=ROUND((G6/H6*100),2)'
Sheet['O7'] = '=ROUND((G7/H7*100),2)'
Sheet['O8'] = '=ROUND((G8/H8*100),2)'
Sheet['O9'] = '=ROUND((G9/H9*100),2)'
Sheet['O10'] = '=ROUND((G10/H10*100),2)'
Sheet['O11'] = '=ROUND((G11/H11*100),2)'
Sheet['O12'] = '=ROUND((G12/H12*100),2)'
Sheet['O13'] = '=ROUND((G13/H13*100),2)'
Sheet['O14'] = '=ROUND((G14/H14*100),2)'
Sheet['O15'] = '=ROUND((G15/H15*100),2)'

Sheet['P4'] = '=ROUND((E4/F4*100),2)'
Sheet['P5'] = '=ROUND((E5/F5*100),2)'
Sheet['P6'] = '=ROUND((E6/F6*100),2)'
Sheet['P7'] = '=ROUND((E7/F7*100),2)'
Sheet['P8'] = '=ROUND((E8/F8*100),2)'
Sheet['P9'] = '=ROUND((E9/F9*100),2)'
Sheet['P10'] = '=ROUND((E10/F10*100),2)'
Sheet['P11'] = '=ROUND((E11/F11*100),2)'
Sheet['P12'] = '=ROUND((E12/F12*100),2)'
Sheet['P13'] = '=ROUND((E13/F13*100),2)'
Sheet['P14'] = '=ROUND((E14/F14*100),2)'
Sheet['P15'] = '=ROUND((E15/F15*100),2)'

Sheet1['T2'] = '=ROUND(((R2+S2)/Q2*100),2)'
Sheet1['T3'] = '=ROUND(((R3+S3)/Q3*100),2)'
Sheet1['T4'] = '=ROUND(((R4+S4)/Q4*100),2)'
Sheet1['T5'] = '=ROUND(((R5+S5)/Q5*100),2)'
Sheet1['T6'] = '=ROUND(((R6+S6)/Q6*100),2)'
Sheet1['T7'] = '=ROUND(((R7+S7)/Q7*100),2)'
Sheet1['T8'] = '=ROUND(((R8+S8)/Q8*100),2)'
Sheet1['T9'] = '=ROUND(((R9+S9)/Q9*100),2)'
Sheet1['T10'] = '=ROUND(((R10+S10)/Q10*100),2)'
Sheet1['T11'] = '=ROUND(((R11+S11)/Q11*100),2)'
Sheet1['T12'] = '=ROUND(((R12+S12)/Q12*100),2)'
Sheet1['T13'] = '=ROUND(((R13+S13)/Q13*100),2)'
Sheet1['T14'] = '=ROUND(((R14+S14)/Q14*100),2)'
Sheet1['T15'] = '=ROUND(((R15+S15)/Q15*100),2)'
Sheet1['T16'] = '=ROUND(((R16+S16)/Q16*100),2)'
Sheet1['T17'] = '=ROUND(((R17+S17)/Q17*100),2)'
Sheet1['T18'] = '=ROUND(((R18+S18)/Q18*100),2)'
Sheet1['T19'] = '=ROUND(((R19+S19)/Q19*100),2)'
Sheet1['T20'] = '=ROUND(((R20+S20)/Q20*100),2)'
Sheet1['T21'] = '=ROUND(((R21+S21)/Q21*100),2)'
Sheet1['T22'] = '=ROUND(((R22+S22)/Q22*100),2)'
Sheet1['T23'] = '=ROUND(((R23+S23)/Q23*100),2)'
Sheet1['T24'] = '=ROUND(((R24+S24)/Q24*100),2)'
Sheet1['T25'] = '=ROUND(((R25+S25)/Q25*100),2)'
Sheet1['T26'] = '=ROUND(((R26+S26)/Q26*100),2)'
Sheet1['T27'] = '=ROUND(((R27+S27)/Q27*100),2)'
Sheet1['T28'] = '=ROUND(((R28+S28)/Q28*100),2)'
Sheet1['T29'] = '=ROUND(((R29+S29)/Q29*100),2)'
Sheet1['T30'] = '=ROUND(((R30+S30)/Q30*100),2)'
Sheet1['T31'] = '=ROUND(((R31+S31)/Q31*100),2)'
Sheet1['T32'] = '=ROUND(((R32+S32)/Q32*100),2)'
Sheet1['T33'] = '=ROUND(((R33+S33)/Q33*100),2)'
Sheet1['T34'] = '=ROUND(((R34+S34)/Q34*100),2)'
Sheet1['T35'] = '=ROUND(((R35+S35)/Q35*100),2)'
Sheet1['T36'] = '=ROUND(((R36+S36)/Q36*100),2)'
Sheet1['T37'] = '=ROUND(((R37+S37)/Q37*100),2)'
Sheet1['T38'] = '=ROUND(((R38+S38)/Q38*100),2)'
Sheet1['T39'] = '=ROUND(((R39+S39)/Q39*100),2)'
Sheet1['T40'] = '=ROUND(((R40+S40)/Q40*100),2)'
Sheet1['T41'] = '=ROUND(((R41+S41)/Q41*100),2)'
Sheet1['T42'] = '=ROUND(((R42+S42)/Q42*100),2)'
Sheet1['T43'] = '=ROUND(((R43+S43)/Q43*100),2)'
Sheet1['T44'] = '=ROUND(((R44+S44)/Q44*100),2)'
Sheet1['T45'] = '=ROUND(((R45+S45)/Q45*100),2)'
Sheet1['T46'] = '=ROUND(((R46+S46)/Q46*100),2)'
Sheet1['T47'] = '=ROUND(((R47+S47)/Q47*100),2)'
Sheet1['T48'] = '=ROUND(((R48+S48)/Q48*100),2)'
Sheet1['T49'] = '=ROUND(((R49+S49)/Q49*100),2)'
Sheet1['T50'] = '=ROUND(((R50+S50)/Q50*100),2)'
Sheet1['T51'] = '=ROUND(((R51+S51)/Q51*100),2)'
Sheet1['T52'] = '=ROUND(((R52+S52)/Q52*100),2)'
Sheet1['T53'] = '=ROUND(((R53+S53)/Q53*100),2)'
Sheet1['T54'] = '=ROUND(((R54+S54)/Q54*100),2)'
Sheet1['T55'] = '=ROUND(((R55+S55)/Q55*100),2)'
Sheet1['T56'] = '=ROUND(((R56+S56)/Q56*100),2)'
Sheet1['T57'] = '=ROUND(((R57+S57)/Q57*100),2)'
Sheet1['T58'] = '=ROUND(((R58+S58)/Q58*100),2)'
Sheet1['T59'] = '=ROUND(((R59+S59)/Q59*100),2)'
Sheet1['T60'] = '=ROUND(((R60+S60)/Q60*100),2)'
Sheet1['T61'] = '=ROUND(((R61+S61)/Q61*100),2)'
Sheet1['T62'] = '=ROUND(((R62+S62)/Q62*100),2)'
Sheet1['T63'] = '=ROUND(((R63+S63)/Q63*100),2)'
Sheet1['T64'] = '=ROUND(((R64+S64)/Q64*100),2)'
Sheet1['T65'] = '=ROUND(((R65+S65)/Q65*100),2)'
Sheet1['T66'] = '=ROUND(((R66+S66)/Q66*100),2)'
Sheet1['T67'] = '=ROUND(((R67+S67)/Q67*100),2)'
Sheet1['T68'] = '=ROUND(((R68+S68)/Q68*100),2)'
Sheet1['T69'] = '=ROUND(((R69+S69)/Q69*100),2)'
Sheet1['T70'] = '=ROUND(((R70+S70)/Q70*100),2)'
Sheet1['T71'] = '=ROUND(((R71+S71)/Q71*100),2)'
Sheet1['T72'] = '=ROUND(((R73+S72)/Q72*100),2)'
Sheet1['T73'] = '=ROUND(((R73+S73)/Q73*100),2)'
Sheet1['T74'] = '=ROUND(((R74+S74)/Q74*100),2)'
Sheet1['T75'] = '=ROUND(((R75+S75)/Q75*100),2)'
Sheet1['T76'] = '=ROUND(((R76+S76)/Q76*100),2)'
Sheet1['T77'] = '=ROUND(((R77+S77)/Q77*100),2)'
Sheet1['T78'] = '=ROUND(((R78+S78)/Q78*100),2)'
Sheet1['T79'] = '=ROUND(((R79+S79)/Q79*100),2)'
Sheet1['T80'] = '=ROUND(((R80+S80)/Q80*100),2)'
Sheet1['T81'] = '=ROUND(((R81+S81)/Q81*100),2)'
Sheet1['T82'] = '=ROUND(((R83+S82)/Q82*100),2)'
Sheet1['T83'] = '=ROUND(((R83+S83)/Q83*100),2)'
Sheet1['T84'] = '=ROUND(((R84+S84)/Q84*100),2)'
Sheet1['T85'] = '=ROUND(((R85+S85)/Q85*100),2)'
Sheet1['T86'] = '=ROUND(((R86+S86)/Q86*100),2)'
Sheet1['T87'] = '=ROUND(((R87+S87)/Q87*100),2)'
Sheet1['T88'] = '=ROUND(((R88+S88)/Q88*100),2)'
Sheet1['T89'] = '=ROUND(((R89+S89)/Q89*100),2)'
Sheet1['T90'] = '=ROUND(((R90+S90)/Q90*100),2)'

Sheet1['U2'] = '=ROUND((K2/L2*100),2)'
Sheet1['U3'] = '=ROUND((K3/L3*100),2)'
Sheet1['U4'] = '=ROUND((K4/L4*100),2)'
Sheet1['U5'] = '=ROUND((K5/L5*100),2)'
Sheet1['U6'] = '=ROUND((K6/L6*100),2)'
Sheet1['U7'] = '=ROUND((K7/L7*100),2)'
Sheet1['U8'] = '=ROUND((K8/L8*100),2)'
Sheet1['U9'] = '=ROUND((K9/L9*100),2)'
Sheet1['U10'] = '=ROUND((K10/L10*100),2)'
Sheet1['U11'] = '=ROUND((K11/L11*100),2)'
Sheet1['U12'] = '=ROUND((K12/L12*100),2)'
Sheet1['U13'] = '=ROUND((K13/L13*100),2)'
Sheet1['U14'] = '=ROUND((K14/L14*100),2)'
Sheet1['U15'] = '=ROUND((K15/L15*100),2)'
Sheet1['U16'] = '=ROUND((K16/L16*100),2)'
Sheet1['U17'] = '=ROUND((K17/L17*100),2)'
Sheet1['U18'] = '=ROUND((K18/L18*100),2)'
Sheet1['U19'] = '=ROUND((K19/L19*100),2)'
Sheet1['U20'] = '=ROUND((K20/L20*100),2)'
Sheet1['U21'] = '=ROUND((K21/L21*100),2)'
Sheet1['U22'] = '=ROUND((K22/L22*100),2)'
Sheet1['U23'] = '=ROUND((K23/L23*100),2)'
Sheet1['U24'] = '=ROUND((K24/L24*100),2)'
Sheet1['U25'] = '=ROUND((K25/L25*100),2)'
Sheet1['U26'] = '=ROUND((K26/L26*100),2)'
Sheet1['U27'] = '=ROUND((K27/L27*100),2)'
Sheet1['U28'] = '=ROUND((K28/L28*100),2)'
Sheet1['U29'] = '=ROUND((K29/L29*100),2)'
Sheet1['U30'] = '=ROUND((K30/L30*100),2)'
Sheet1['U31'] = '=ROUND((K31/L31*100),2)'
Sheet1['U32'] = '=ROUND((K32/L32*100),2)'
Sheet1['U33'] = '=ROUND((K33/L33*100),2)'
Sheet1['U34'] = '=ROUND((K34/L34*100),2)'
Sheet1['U35'] = '=ROUND((K35/L35*100),2)'
Sheet1['U36'] = '=ROUND((K36/L36*100),2)'
Sheet1['U37'] = '=ROUND((K37/L37*100),2)'
Sheet1['U38'] = '=ROUND((K38/L38*100),2)'
Sheet1['U39'] = '=ROUND((K39/L39*100),2)'
Sheet1['U40'] = '=ROUND((K40/L40*100),2)'
Sheet1['U41'] = '=ROUND((K41/L41*100),2)'
Sheet1['U42'] = '=ROUND((K42/L42*100),2)'
Sheet1['U43'] = '=ROUND((K43/L43*100),2)'
Sheet1['U44'] = '=ROUND((K44/L44*100),2)'
Sheet1['U45'] = '=ROUND((K45/L45*100),2)'
Sheet1['U46'] = '=ROUND((K46/L46*100),2)'
Sheet1['U47'] = '=ROUND((K47/L47*100),2)'
Sheet1['U48'] = '=ROUND((K48/L48*100),2)'
Sheet1['U49'] = '=ROUND((K49/L49*100),2)'
Sheet1['U50'] = '=ROUND((K50/L50*100),2)'
Sheet1['U51'] = '=ROUND((K51/L51*100),2)'
Sheet1['U52'] = '=ROUND((K52/L52*100),2)'
Sheet1['U53'] = '=ROUND((K53/L53*100),2)'
Sheet1['U54'] = '=ROUND((K54/L54*100),2)'
Sheet1['U55'] = '=ROUND((K55/L55*100),2)'
Sheet1['U56'] = '=ROUND((K56/L56*100),2)'
Sheet1['U57'] = '=ROUND((K57/L57*100),2)'
Sheet1['U58'] = '=ROUND((K58/L58*100),2)'
Sheet1['U59'] = '=ROUND((K59/L59*100),2)'
Sheet1['U60'] = '=ROUND((K60/L60*100),2)'
Sheet1['U61'] = '=ROUND((K61/L61*100),2)'
Sheet1['U62'] = '=ROUND((K62/L62*100),2)'
Sheet1['U63'] = '=ROUND((K63/L63*100),2)'
Sheet1['U64'] = '=ROUND((K64/L64*100),2)'
Sheet1['U65'] = '=ROUND((K65/L65*100),2)'
Sheet1['U66'] = '=ROUND((K66/L66*100),2)'
Sheet1['U67'] = '=ROUND((K67/L67*100),2)'
Sheet1['U68'] = '=ROUND((K68/L68*100),2)'
Sheet1['U69'] = '=ROUND((K69/L69*100),2)'
Sheet1['U70'] = '=ROUND((K70/L70*100),2)'
Sheet1['U71'] = '=ROUND((K71/L71*100),2)'
Sheet1['U72'] = '=ROUND((K72/L72*100),2)'
Sheet1['U73'] = '=ROUND((K73/L73*100),2)'
Sheet1['U74'] = '=ROUND((K74/L74*100),2)'
Sheet1['U75'] = '=ROUND((K75/L75*100),2)'
Sheet1['U76'] = '=ROUND((K76/L76*100),2)'
Sheet1['U77'] = '=ROUND((K77/L77*100),2)'
Sheet1['U78'] = '=ROUND((K78/L78*100),2)'
Sheet1['U79'] = '=ROUND((K79/L79*100),2)'
Sheet1['U80'] = '=ROUND((K80/L80*100),2)'
Sheet1['U81'] = '=ROUND((K81/L81*100),2)'
Sheet1['U82'] = '=ROUND((K82/L82*100),2)'
Sheet1['U83'] = '=ROUND((K83/L83*100),2)'
Sheet1['U84'] = '=ROUND((K84/L84*100),2)'
Sheet1['U85'] = '=ROUND((K85/L85*100),2)'
Sheet1['U86'] = '=ROUND((K86/L86*100),2)'
Sheet1['U87'] = '=ROUND((K87/L87*100),2)'
Sheet1['U88'] = '=ROUND((K88/L88*100),2)'
Sheet1['U89'] = '=ROUND((K89/L89*100),2)'
Sheet1['U90'] = '=ROUND((K90/L90*100),2)'

Sheet1['V2'] = '=ROUND((M2/N2*100),2)'
Sheet1['V3'] = '=ROUND((M3/N3*100),2)'
Sheet1['V4'] = '=ROUND((M4/N4*100),2)'
Sheet1['V5'] = '=ROUND((M5/N5*100),2)'
Sheet1['V6'] = '=ROUND((M6/N6*100),2)'
Sheet1['V7'] = '=ROUND((M7/N7*100),2)'
Sheet1['V8'] = '=ROUND((M8/N8*100),2)'
Sheet1['V9'] = '=ROUND((M9/N9*100),2)'
Sheet1['V10'] = '=ROUND((M10/N10*100),2)'
Sheet1['V11'] = '=ROUND((M11/N11*100),2)'
Sheet1['V12'] = '=ROUND((M12/N12*100),2)'
Sheet1['V13'] = '=ROUND((M13/N13*100),2)'
Sheet1['V14'] = '=ROUND((M14/N14*100),2)'
Sheet1['V15'] = '=ROUND((M15/N15*100),2)'
Sheet1['V16'] = '=ROUND((M16/N16*100),2)'
Sheet1['V17'] = '=ROUND((M17/N17*100),2)'
Sheet1['V18'] = '=ROUND((M18/N18*100),2)'
Sheet1['V19'] = '=ROUND((M19/N19*100),2)'
Sheet1['V20'] = '=ROUND((M20/N20*100),2)'
Sheet1['V21'] = '=ROUND((M21/N21*100),2)'
Sheet1['V22'] = '=ROUND((M22/N22*100),2)'
Sheet1['V23'] = '=ROUND((M23/N23*100),2)'
Sheet1['V24'] = '=ROUND((M24/N24*100),2)'
Sheet1['V25'] = '=ROUND((M25/N25*100),2)'
Sheet1['V26'] = '=ROUND((M26/N26*100),2)'
Sheet1['V27'] = '=ROUND((M27/N27*100),2)'
Sheet1['V28'] = '=ROUND((M28/N28*100),2)'
Sheet1['V29'] = '=ROUND((M29/N29*100),2)'
Sheet1['V30'] = '=ROUND((M30/N30*100),2)'
Sheet1['V31'] = '=ROUND((M31/N31*100),2)'
Sheet1['V32'] = '=ROUND((M32/N32*100),2)'
Sheet1['V33'] = '=ROUND((M33/N33*100),2)'
Sheet1['V34'] = '=ROUND((M34/N34*100),2)'
Sheet1['V35'] = '=ROUND((M35/N35*100),2)'
Sheet1['V36'] = '=ROUND((M36/N36*100),2)'
Sheet1['V37'] = '=ROUND((M37/N37*100),2)'
Sheet1['V38'] = '=ROUND((M38/N38*100),2)'
Sheet1['V39'] = '=ROUND((M39/N39*100),2)'
Sheet1['V40'] = '=ROUND((M40/N40*100),2)'
Sheet1['V41'] = '=ROUND((M41/N41*100),2)'
Sheet1['V42'] = '=ROUND((M42/N42*100),2)'
Sheet1['V43'] = '=ROUND((M43/N43*100),2)'
Sheet1['V44'] = '=ROUND((M44/N44*100),2)'
Sheet1['V45'] = '=ROUND((M45/N45*100),2)'
Sheet1['V46'] = '=ROUND((M46/N46*100),2)'
Sheet1['V47'] = '=ROUND((M47/N47*100),2)'
Sheet1['V48'] = '=ROUND((M48/N48*100),2)'
Sheet1['V49'] = '=ROUND((M49/N49*100),2)'
Sheet1['V50'] = '=ROUND((M50/N50*100),2)'
Sheet1['V51'] = '=ROUND((M51/N51*100),2)'
Sheet1['V52'] = '=ROUND((M52/N52*100),2)'
Sheet1['V53'] = '=ROUND((M53/N53*100),2)'
Sheet1['V54'] = '=ROUND((M54/N54*100),2)'
Sheet1['V55'] = '=ROUND((M55/N55*100),2)'
Sheet1['V56'] = '=ROUND((M56/N56*100),2)'
Sheet1['V57'] = '=ROUND((M57/N57*100),2)'
Sheet1['V58'] = '=ROUND((M58/N58*100),2)'
Sheet1['V59'] = '=ROUND((M59/N59*100),2)'
Sheet1['V60'] = '=ROUND((M60/N60*100),2)'
Sheet1['V60'] = '=ROUND((M60/N60*100),2)'
Sheet1['V61'] = '=ROUND((M61/N61*100),2)'
Sheet1['V62'] = '=ROUND((M62/N62*100),2)'
Sheet1['V63'] = '=ROUND((M63/N63*100),2)'
Sheet1['V64'] = '=ROUND((M64/N64*100),2)'
Sheet1['V66'] = '=ROUND((M66/N66*100),2)'
Sheet1['V65'] = '=ROUND((M65/N65*100),2)'
Sheet1['V67'] = '=ROUND((M67/N67*100),2)'
Sheet1['V68'] = '=ROUND((M68/N68*100),2)'
Sheet1['V69'] = '=ROUND((M69/N69*100),2)'
Sheet1['V70'] = '=ROUND((M70/N70*100),2)'
Sheet1['V71'] = '=ROUND((M71/N71*100),2)'
Sheet1['V72'] = '=ROUND((M72/N72*100),2)'
Sheet1['V73'] = '=ROUND((M73/N73*100),2)'
Sheet1['V74'] = '=ROUND((M74/N74*100),2)'
Sheet1['V76'] = '=ROUND((M76/N76*100),2)'
Sheet1['V75'] = '=ROUND((M75/N75*100),2)'
Sheet1['V77'] = '=ROUND((M77/N77*100),2)'
Sheet1['V78'] = '=ROUND((M78/N78*100),2)'
Sheet1['V79'] = '=ROUND((M79/N79*100),2)'
Sheet1['V80'] = '=ROUND((M80/N80*100),2)'
Sheet1['V81'] = '=ROUND((M81/N81*100),2)'
Sheet1['V82'] = '=ROUND((M82/N82*100),2)'
Sheet1['V83'] = '=ROUND((M83/N83*100),2)'
Sheet1['V84'] = '=ROUND((M84/N84*100),2)'
Sheet1['V86'] = '=ROUND((M86/N86*100),2)'
Sheet1['V85'] = '=ROUND((M85/N85*100),2)'
Sheet1['V87'] = '=ROUND((M87/N87*100),2)'
Sheet1['V88'] = '=ROUND((M88/N88*100),2)'
Sheet1['V89'] = '=ROUND((M89/N89*100),2)'
Sheet1['V90'] = '=ROUND((M90/N90*100),2)'

# Sheet.move_range("A1:Q15", rows=2, translate=True)
Sheet1.move_range("A1:Z100", rows=2, translate=True)

Sheet['A1'] = 'Date'
Sheet['B1'] = date_str
Sheet['H1'] = 'Indian designs Exports Pvt Ltd'
Sheet1['A1'] = 'Date'
Sheet1['B1'] = date_str
Sheet1['H1'] = 'Indian designs Exports Pvt Ltd'
Sheet1['A3'] = 'Sl No.'

Sheet['A3'] = 'Sl No.'
Sheet['Q3'] = 'Running Batches Count'
Sheet['Q4'] = '=COUNTIF(Sheet1!B4:B100,B4)'
Sheet['Q5'] = '=COUNTIF(Sheet1!B4:B100,B5)'
Sheet['Q6'] = '=COUNTIF(Sheet1!B4:B100,B6)'
Sheet['Q7'] = '=COUNTIF(Sheet1!B4:B100,B7)'
Sheet['Q8'] = '=COUNTIF(Sheet1!B4:B100,B8)'
Sheet['Q9'] = '=COUNTIF(Sheet1!B4:B100,B9)'
Sheet['Q10'] = '=COUNTIF(Sheet1!B4:B100,B10)'
Sheet['Q11'] = '=COUNTIF(Sheet1!B4:B100,B11)'
Sheet['Q12'] = '=COUNTIF(Sheet1!B4:B100,B12)'
Sheet['Q13'] = '=COUNTIF(Sheet1!B4:B100,B13)'
Sheet['Q14'] = '=COUNTIF(Sheet1!B4:B100,B14)'
Sheet['Q15'] = '=COUNTIF(Sheet1!B4:B100,B15)'

Sheet['C16'] = '=SUM(C4:C15)'
Sheet['D16'] = '=SUM(D4:D15)'
Sheet['E16'] = '=SUM(E4:E15)'
Sheet['F16'] = '=SUM(F4:F15)'
Sheet['G16'] = '=SUM(G4:G15)'
Sheet['H16'] = '=SUM(H4:H15)'
Sheet['I16'] = '=SUM(I4:I15)'
Sheet['J16'] = '=SUM(J4:J15)'
Sheet['K16'] = '=SUM(K4:K15)'
Sheet['L16'] = '=SUM(L4:L15)'
Sheet['M16'] = '=SUM(M4:M15)'
Sheet['N16'] = '=ROUND(((L16+M16)/K16*100),2)'
Sheet['P16'] = '=ROUND((G16/H16*100),2)'
Sheet['O16'] = '=ROUND((E16/F16*100),2)'
Sheet['Q16'] = '=SUM(Q4:Q15)'
Sheet['A16'] = 'Summary'
Sheet.merge_cells('A16:B16') 

Sheet['A1'].font = Font(size=15, underline='single', color='0000FF', bold=True, italic=True)
Sheet['B1'].font = Font(size=15, underline='single', color='0000FF', bold=True, italic=True)
Sheet['H1'].font = Font(size=15, underline='single', color='0000FF', bold=True, italic=True)
Sheet1['A1'].font = Font(size=15, underline='single', color='0000FF', bold=True, italic=True)
Sheet1['B1'].font = Font(size=15, underline='single', color='0000FF', bold=True, italic=True)
Sheet1['H1'].font = Font(size=15, underline='single', color='0000FF', bold=True, italic=True)

Sheet['A3'].font = Font(size=12, bold=True)
Sheet['B3'].font = Font(size=12, bold=True)
Sheet['C3'].font = Font(size=12, bold=True)
Sheet['D3'].font = Font(size=12, bold=True)
Sheet['E3'].font = Font(size=12, bold=True)
Sheet['F3'].font = Font(size=12, bold=True)
Sheet['G3'].font = Font(size=12, bold=True)
Sheet['H3'].font = Font(size=12, bold=True)
Sheet['I3'].font = Font(size=12, bold=True)
Sheet['J3'].font = Font(size=12, bold=True)
Sheet['K3'].font = Font(size=12, bold=True)
Sheet['L3'].font = Font(size=12, bold=True)
Sheet['M3'].font = Font(size=12, bold=True)
Sheet['N3'].font = Font(size=12, bold=True)
Sheet['O3'].font = Font(size=12, bold=True)
Sheet['P3'].font = Font(size=12, bold=True)
Sheet['Q3'].font = Font(size=12, bold=True)

Sheet1['A3'].font = Font(size=12, bold=True)
Sheet1['B3'].font = Font(size=12, bold=True)
Sheet1['C3'].font = Font(size=12, bold=True)
Sheet1['D3'].font = Font(size=12, bold=True)
Sheet1['E3'].font = Font(size=12, bold=True)
Sheet1['F3'].font = Font(size=12, bold=True)
Sheet1['G3'].font = Font(size=12, bold=True)
Sheet1['H3'].font = Font(size=12, bold=True)
Sheet1['I3'].font = Font(size=12, bold=True)
Sheet1['J3'].font = Font(size=12, bold=True)
Sheet1['K3'].font = Font(size=12, bold=True)
Sheet1['L3'].font = Font(size=12, bold=True)
Sheet1['M3'].font = Font(size=12, bold=True)
Sheet1['N3'].font = Font(size=12, bold=True)
Sheet1['O3'].font = Font(size=12, bold=True)
Sheet1['P3'].font = Font(size=12, bold=True)
Sheet1['Q3'].font = Font(size=12, bold=True)
Sheet1['R3'].font = Font(size=12, bold=True)
Sheet1['S3'].font = Font(size=12, bold=True)
Sheet1['T3'].font = Font(size=12, bold=True)
Sheet1['U3'].font = Font(size=12, bold=True)
Sheet1['V3'].font = Font(size=12, bold=True)

Sheet['A16'].font = Font(size=12, bold=True)
Sheet['B16'].font = Font(size=12, bold=True)
Sheet['C16'].font = Font(size=12, bold=True)
Sheet['D16'].font = Font(size=12, bold=True)
Sheet['E16'].font = Font(size=12, bold=True)
Sheet['F16'].font = Font(size=12, bold=True)
Sheet['G16'].font = Font(size=12, bold=True)
Sheet['H16'].font = Font(size=12, bold=True)
Sheet['I16'].font = Font(size=12, bold=True)
Sheet['J16'].font = Font(size=12, bold=True)
Sheet['K16'].font = Font(size=12, bold=True)
Sheet['L16'].font = Font(size=12, bold=True)
Sheet['M16'].font = Font(size=12, bold=True)
Sheet['N16'].font = Font(size=12, bold=True)
Sheet['O16'].font = Font(size=12, bold=True)
Sheet['P16'].font = Font(size=12, bold=True)
Sheet['Q16'].font = Font(size=12, bold=True)

Sheet['A3'].fill = PatternFill(fill_type='solid', start_color='CECECE')
Sheet['B3'].fill = PatternFill(fill_type='solid', start_color='CECECE')
Sheet['C3'].fill = PatternFill(fill_type='solid', start_color='CECECE')
Sheet['D3'].fill = PatternFill(fill_type='solid', start_color='CECECE')
Sheet['E3'].fill = PatternFill(fill_type='solid', start_color='CECECE')
Sheet['F3'].fill = PatternFill(fill_type='solid', start_color='CECECE')
Sheet['G3'].fill = PatternFill(fill_type='solid', start_color='CECECE')
Sheet['H3'].fill = PatternFill(fill_type='solid', start_color='CECECE')
Sheet['I3'].fill = PatternFill(fill_type='solid', start_color='CECECE')
Sheet['J3'].fill = PatternFill(fill_type='solid', start_color='CECECE')
Sheet['K3'].fill = PatternFill(fill_type='solid', start_color='CECECE')
Sheet['L3'].fill = PatternFill(fill_type='solid', start_color='CECECE')
Sheet['M3'].fill = PatternFill(fill_type='solid', start_color='CECECE')
Sheet['N3'].fill = PatternFill(fill_type='solid', start_color='CECECE')
Sheet['O3'].fill = PatternFill(fill_type='solid', start_color='CECECE')
Sheet['P3'].fill = PatternFill(fill_type='solid', start_color='CECECE')
Sheet['Q3'].fill = PatternFill(fill_type='solid', start_color='CECECE')
Sheet1['A3'].fill = PatternFill(fill_type='solid', start_color='CECECE')
Sheet1['B3'].fill = PatternFill(fill_type='solid', start_color='CECECE')
Sheet1['C3'].fill = PatternFill(fill_type='solid', start_color='CECECE')
Sheet1['D3'].fill = PatternFill(fill_type='solid', start_color='CECECE')
Sheet1['E3'].fill = PatternFill(fill_type='solid', start_color='CECECE')
Sheet1['F3'].fill = PatternFill(fill_type='solid', start_color='CECECE')
Sheet1['G3'].fill = PatternFill(fill_type='solid', start_color='CECECE')
Sheet1['H3'].fill = PatternFill(fill_type='solid', start_color='CECECE')
Sheet1['I3'].fill = PatternFill(fill_type='solid', start_color='CECECE')
Sheet1['J3'].fill = PatternFill(fill_type='solid', start_color='CECECE')
Sheet1['K3'].fill = PatternFill(fill_type='solid', start_color='CECECE')
Sheet1['L3'].fill = PatternFill(fill_type='solid', start_color='CECECE')
Sheet1['M3'].fill = PatternFill(fill_type='solid', start_color='CECECE')
Sheet1['N3'].fill = PatternFill(fill_type='solid', start_color='CECECE')
Sheet1['O3'].fill = PatternFill(fill_type='solid', start_color='CECECE')
Sheet1['P3'].fill = PatternFill(fill_type='solid', start_color='CECECE')
Sheet1['Q3'].fill = PatternFill(fill_type='solid', start_color='CECECE')
Sheet1['R3'].fill = PatternFill(fill_type='solid', start_color='CECECE')
Sheet1['S3'].fill = PatternFill(fill_type='solid', start_color='CECECE')
Sheet1['T3'].fill = PatternFill(fill_type='solid', start_color='CECECE')
Sheet1['U3'].fill = PatternFill(fill_type='solid', start_color='CECECE')
Sheet1['V3'].fill = PatternFill(fill_type='solid', start_color='CECECE')

Sheet['A16'].fill = PatternFill(fill_type='solid', start_color='CECECE')
Sheet['B16'].fill = PatternFill(fill_type='solid', start_color='CECECE')
Sheet['C16'].fill = PatternFill(fill_type='solid', start_color='CECECE')
Sheet['D16'].fill = PatternFill(fill_type='solid', start_color='CECECE')
Sheet['E16'].fill = PatternFill(fill_type='solid', start_color='CECECE')
Sheet['F16'].fill = PatternFill(fill_type='solid', start_color='CECECE')
Sheet['G16'].fill = PatternFill(fill_type='solid', start_color='CECECE')
Sheet['H16'].fill = PatternFill(fill_type='solid', start_color='CECECE')
Sheet['I16'].fill = PatternFill(fill_type='solid', start_color='CECECE')
Sheet['J16'].fill = PatternFill(fill_type='solid', start_color='CECECE')
Sheet['K16'].fill = PatternFill(fill_type='solid', start_color='CECECE')
Sheet['L16'].fill = PatternFill(fill_type='solid', start_color='CECECE')
Sheet['M16'].fill = PatternFill(fill_type='solid', start_color='CECECE')
Sheet['N16'].fill = PatternFill(fill_type='solid', start_color='CECECE')
Sheet['P16'].fill = PatternFill(fill_type='solid', start_color='CECECE')
Sheet['O16'].fill = PatternFill(fill_type='solid', start_color='CECECE')
Sheet['Q16'].fill = PatternFill(fill_type='solid', start_color='CECECE')

for row in Sheet.iter_rows():
    for cell in row:
        cell.alignment = openpyxl.styles.Alignment(horizontal='center')

for row in range:
    for cell in row:
        cell.border = openpyxl.styles.borders.Border(
            left=openpyxl.styles.borders.Side(style='thin'),
            right=openpyxl.styles.borders.Side(style='thin'),
            top=openpyxl.styles.borders.Side(style='thin'),
            bottom=openpyxl.styles.borders.Side(style='thin')
        )
         

wb.save(xlsx_file_name)
