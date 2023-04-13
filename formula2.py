import openpyxl
from datetime import datetime,timedelta
from datetime import date


date_today  = date.today() 
date_str = datetime.now().strftime('%Y-%m-%d')
date_input = datetime.now().strftime('%Y%m%d')
json_file_name = 'file-'+date_str+'.json'
xlsx_file_name =  'file-'+date_str+'.xlsx'

# Load the Excel workbook
wb = openpyxl.load_workbook(xlsx_file_name)

# Select the sheet and the column you want to apply the formula to
sheet = wb['Sheet1']
column = 'P'

# Get the last row of data in the column
last_row = sheet.max_row

# Iterate over the rows in the column
for row in range(4, last_row + 1):
    # Apply the formula to the cell in column P, using relative cell references
    sheet.cell(row=row, column=16).value = '=ROUND(((N' + str(row) + '+O' + str(row) + ')/M' + str(row) + '*100),2)'

# Save the workbook
wb.save(xlsx_file_name)